from datetime import datetime
from typing import Self
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

import pandas as pd

from config import ConfigManager

class EmployeePerformanceController:
    def __init__(self):

        self.status_fills = {
            'S Perfect Record': PatternFill(start_color='A7C7E7', end_color='A7C7E7', fill_type='solid'),
            'Good Record': PatternFill(start_color='C6DBF0', end_color='C6DBF0', fill_type='solid'),
            'Needs Improvement': PatternFill(start_color='FFF2A6', end_color='FFF2A6', fill_type='solid'),
            'At Risk': PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid'),
            'Recommended for Dismissal': PatternFill(start_color='F4A6A6', end_color='F4A6A6', fill_type='solid'),
            'no data': PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        }

        self.trend_fills = {
                'improving': PatternFill(start_color='A8E6CF', end_color='A8E6CF', fill_type='solid'),
                'declining': PatternFill(start_color='F7A1C4', end_color='F7A1C4', fill_type='solid'),
                'stable':    PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid'),
        }
        self.red_fill = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        self.yellow_fill = PatternFill(start_color='FFF9CC', end_color='FFF9CC', fill_type='solid')

        self.odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.even_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        CM = ConfigManager()
        self.XMonth = CM.config['data']["absence-X-M"] | CM.config['data']["target-X-M"]

        return

    def ProcessSummaryMonth(AbsenData, TargetData):
        SummaryM = pd.merge(
            AbsenData,
            TargetData,
            on=['No.Absen', 'Nama', 'Bagian', 'tahun'],
            how='outer'
        )

        Column_M = [
            'No.Absen', 'Nama', 'Bagian', 'tahun', 'branch', 'overall_status_A', 'overall_status_T', 'recent_trend_A',  'recent_trend_T'
        ]

        for i in range(1, 13):
            Column_M.append(f'Total_Absen_B{i}')
            Column_M.append(f'Nilai_Absen_B{i}')
            Column_M.append(f'{i}')

        Column_M =  [col for col in Column_M if col in SummaryM.columns]
        Column_M.append('Keterangan')

        SummaryM = SummaryM[Column_M]
        SummaryM = (
            SummaryM.groupby(['No.Absen'], as_index=False)
                  .agg(lambda x: x.ffill().bfill().iloc[0]) )
        SummaryM = SummaryM.infer_objects(copy=False)
        SummaryM = SummaryM.drop_duplicates(subset='No.Absen', keep='last').reset_index(drop=True)

        # top down
        record = []
        for _, row in SummaryM.iterrows():
            absence = {
                'No.Absen': row['No.Absen'],
                'Nama': row['Nama'],
                'Bagian': row['Bagian'],
                'tahun': row['tahun'],
                'cabang': row['branch'],
                '#': 'absence',
                'Status keseluruhan': row['overall_status_A'],
                'Tren terakhir': row['recent_trend_A']
            }
            target = {
                'No.Absen': row['No.Absen'],
                'Nama': row['Nama'],
                'Bagian': row['Bagian'],
                'tahun': row['tahun'],
                'cabang': row['branch'],
                '#': 'target',
                'Status keseluruhan': row['overall_status_T'],
                'Tren terakhir': row['recent_trend_T']
            }

            for i in range(1, 13):
                absence[f'Total Absen {i}'] = row.get(f'Total_Absen_B{i}', '-')
                absence[f'Nilai bulan {i}'] = row.get(f'Nilai_Absen_B{i}', '-')

                target[f'Total Absen {i}'] = row.get(f'Total_Absen_B{i}', '-')
                target[f'Nilai bulan {i}'] = row.get(str(i), '-')

            absence['Keterangan'] = row.get('Keterangan', '')
            target['Keterangan'] = row.get('Keterangan', '')
            record.append(absence)
            record.append(target)

        result = pd.DataFrame(record)

        return result

    def ProcessSummaryXMonth(self, AbsenXData, TargetXData):
        SummaryXM = pd.merge(
            AbsenXData,
            TargetXData,
            on=['No.Absen', 'Nama', 'Bagian', 'tahun'],
            how='outer'
        )
        Column_XM = [
            'No.Absen', 'Nama', 'Bagian', 'tahun', 'branch', 'recent_status_A', 'recent_status_T', 'recent_trend_A',  'recent_trend_T'
        ]
        current_month = datetime.now().month
        last_x_months = [(current_month - i-1) % 12 + 1 for i in range(self.XMonth)]
        last_x_months.sort()

        for m in last_x_months:
            Column_XM.append(f'Total_Absen_B{m}')
            Column_XM.append(f'Nilai_Absen_B{m}')
            Column_XM.append(f'{m}')

        Column_XM =  [col for col in Column_XM if col in SummaryXM.columns]
        Column_XM.append('Keterangan')
        SummaryXM = SummaryXM[Column_XM]

        SummaryXM = (
            SummaryXM.groupby(['No.Absen'], as_index=False)
                  .agg(lambda x: x.ffill().bfill().iloc[0]) )
        SummaryXM = SummaryXM.infer_objects(copy=False)
        SummaryXM = SummaryXM.drop_duplicates(subset='No.Absen', keep='last').reset_index(drop=True)

        record = []
        for _, row in SummaryXM.iterrows():
            absence = {
                'No.Absen': row['No.Absen'],
                'Nama': row['Nama'],
                'Bagian': row['Bagian'],
                'tahun': row['tahun'],
                'cabang': row['branch'],
                '#': 'absence',
                'Status keseluruhan': row['overall_status_A'],
                'Tren terakhir': row['recent_trend_A']
            }
            target = {
                'No.Absen': row['No.Absen'],
                'Nama': row['Nama'],
                'Bagian': row['Bagian'],
                'tahun': row['tahun'],
                'cabang': row['branch'],
                '#': 'target',
                'Status keseluruhan': row['overall_status_T'],
                'Tren terakhir': row['recent_trend_T']
            }

            for i in range(1, 13):
                absence[f'Total Absen {i}'] = row.get(f'Total_Absen_B{i}', '-')
                absence[f'Nilai bulan {i}'] = row.get(f'Nilai_Absen_B{i}', '-')

                target[f'Total Absen {i}'] = row.get(f'Total_Absen_B{i}', '-')
                target[f'Nilai bulan {i}'] = row.get(str(i), '-')

            absence['Keterangan'] = row.get('Keterangan', '')
            target['Keterangan'] = row.get('Keterangan', '')
            record.append(absence)
            record.append(target)

        result = pd.DataFrame(record)
        return result

    def setGradeColour(self, ws, Column_M, dataCols, start_row, end_row, colCondition):
        grade_columns = [col for col in Column_M if col.startswith(colCondition)]
        for col_name in grade_columns:
            col_index = dataCols.get_loc(col_name) + 2
            col_letter = get_column_letter(col_index)
            cell_range = f'{col_letter}{start_row}:{col_letter}{end_row}'

            formula_e = f'${col_letter}{start_row}="E"'
            formula_d = f'${col_letter}{start_row}="D"'

            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula_e], fill=self.red_fill))
            ws.conditional_formatting.add(cell_range, FormulaRule(formula=[formula_d], fill=self.yellow_fill))
        return

    def setOverallStatus(self, ws, dataCols, start_row, end_row):
        dataCols=dataCols.astype(str)
        normalized_cols = dataCols.str.replace("_", " ", regex=False)
        found = any("Status keseluruhan" in col for col in normalized_cols) or any("Status terakhir" in col for col in normalized_cols) or any("overall status" in col for col in normalized_cols)
        if found:
            loc =  next(
                (dataCols[i] for i, col in enumerate(normalized_cols) if col.startswith("Status keseluruhan") or col.startswith("Status terakhir") or col.startswith("overall status")),
                ""
            )

            status_index = dataCols.get_loc(loc) + 2
            status_col = get_column_letter(status_index)
            status_range = f'{status_col}{start_row}:{status_col}{end_row}'

            for status, fill in self.status_fills.items():
                formula = f'${status_col}{start_row}="{status}"'
                ws.conditional_formatting.add(status_range, FormulaRule(formula=[formula], fill=fill))
        return

    def setRecentTrend(self, ws, dataCols, start_row, end_row):
        dataCols=dataCols.astype(str)
        normalized_cols = dataCols.str.replace("_", " ", regex=False)
        found = any("Tren terakhir" in col for col in normalized_cols) or any("recent trend" in col for col in normalized_cols)
        if found:
            loc =  next(
                (dataCols[i] for i, col in enumerate(normalized_cols) if col.startswith("Tren terakhir") or col.startswith("recent trend")),
                ""
            )
            trend_index = dataCols.get_loc(loc) + 2
            trend_col = get_column_letter(trend_index)
            trend_range = f'{trend_col}{start_row}:{trend_col}{end_row}'

            for trend, fill in self.trend_fills.items():
                formula = f'${trend_col}{start_row}="{trend}"'
                ws.conditional_formatting.add(trend_range, FormulaRule(formula=[formula], fill=fill))

        return

    def setAlternatingFill(self, ws):
        thin_border = Border(
            left=Side(style='none'),
            right=Side(style='none'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            row_idx = row[0].row
            for cell in row:
                # Apply border
                cell.border = thin_border

                # Fill even or odd row
                if row_idx % 2 == 0:
                    cell.fill = self.even_fill
                else:
                    cell.fill = self.odd_fill
        return

    def formatMetaData(self, w):
        wb = load_workbook(w)
        ws = wb['metadata']
        # Define styles
        header_font = Font(bold=True)
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Apply styles
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
                if cell.row == 1:
                    cell.font = header_font
                    cell.fill = header_fill

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column name
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width
        wb.save(w)
        return

    def formatExcelMonth(self, w, sheetname, data, type):
        wb = load_workbook(w)
        ws = wb[sheetname]

        ws.row_dimensions[1].height = 30
        ws.auto_filter.ref = ws.dimensions

        Column_M = []
        colCondition = ''
        if type == "combine m":
            Column_M = [
                'No.Absen', 'Nama', 'Bagian', 'tahun', 'cabang', '#', 'Status keseluruhan', 'Tren terakhir'
            ]
            colCondition = 'Nilai bulan'
        elif type == "combine xm":
            Column_M = [
                'No.Absen', 'Nama', 'Bagian', 'tahun', 'cabang', '#', 'Status terakhir', 'Tren terakhir'
            ]
            colCondition = 'Nilai bulan'
        elif type == 'Absence':
            Column_M = [
                'No.Absen', 'Nama', 'Bagian', 'branch',  'tahun',  'overall_status_A', 'recent_trend_A'
            ]
            colCondition = 'Nilai_Absen_B'
        elif type == 'Target':
            Column_M = [
                'No.Absen', 'Nama', 'Bagian', 'tahun',  'overall_status_T', 'recent_trend_T'
            ]
            colCondition = ''
        elif type == 'metadata':
            Column_M = [
                'Category, Grade', 'count', 'Metric', 'Value'
            ]
            colCondition = ''

        dataCols = data.columns
        for i in range(1, 13):
            if type == 'combine m':
                Column_M.append(f'Total Absen {i}')
                Column_M.append(f'Nilai bulan {i}')
            elif type == 'combine xm':
                Column_M.append(f'Total Absen {i}')
                Column_M.append(f'Nilai bulan {i}')
            elif type == 'Absence':
                Column_M.append(f'Total_Absen_B{i}')
                Column_M.append(f'Nilai_Absen_B{i}')
            elif type == 'Target':
                Column_M.append(f'{i}')

        Column_M = [col for col in Column_M if col in dataCols]

        if type is not "metadata": Column_M.append('Keterangan')

        start_row = 2
        end_row = ws.max_row

        self.setGradeColour(ws, Column_M, dataCols, start_row, end_row, colCondition)
        self.setOverallStatus(ws, dataCols, start_row, end_row)
        self.setRecentTrend(ws, dataCols, start_row, end_row)
        self.setAlternatingFill(ws)
        # max_row = ws.max_row
        # max_col = ws.max_column

        # ref = f"A1:{get_column_letter(max_col)}{max_row}"
        # table = Table(displayName="DynamicTable", ref=ref)

        # ws.add_table(table)



        wb.save(w)
        return
